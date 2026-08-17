"""Sync daily attendance to Excel sheets (one file per school per date)."""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app import config
from app.class_sort import sort_students
from app.database import db, rows_to_list, today_str


def excel_path_for(school_id: int, school_name: str, date: str | None = None) -> Path:
    date = date or today_str()
    safe_name = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in school_name)
    filename = f"{safe_name}_{school_id}_{date}.xlsx"
    return config.EXPORTS_DIR / filename


def _style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def write_attendance_excel(
    school_id: int,
    school_name: str,
    date: str | None = None,
) -> Path:
    """Rebuild Excel for a school/date from the database."""
    date = date or today_str()
    path = excel_path_for(school_id, school_name, date)

    with db() as conn:
        rows = conn.execute(
            """
            SELECT student_id, name, class_name, time_in, time_out, is_present
            FROM attendance
            WHERE school_id = ? AND date = ?
            """,
            (school_id, date),
        ).fetchall()
        records = rows_to_list(rows)

        # Also include enrolled students with no attendance row yet (absent)
        enrolled = conn.execute(
            """
            SELECT student_id, name, class_name
            FROM students
            WHERE school_id = ? AND is_active = 1
            """,
            (school_id,),
        ).fetchall()
        enrolled_list = rows_to_list(enrolled)

    present_ids = {r["student_id"] for r in records}
    full: list[dict[str, Any]] = list(records)
    for s in enrolled_list:
        if s["student_id"] not in present_ids:
            full.append(
                {
                    "student_id": s["student_id"],
                    "name": s["name"],
                    "class_name": s["class_name"],
                    "time_in": "",
                    "time_out": "",
                    "is_present": 0,
                }
            )

    full = sort_students(full)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    headers = [
        "Student ID",
        "Name",
        "Class",
        "Date",
        "Time In",
        "Time Out",
        "Is Present",
    ]
    ws.append(headers)
    _style_header(ws)

    for r in full:
        present = "Yes" if r.get("is_present") else "No"
        ws.append(
            [
                r.get("student_id", ""),
                r.get("name", ""),
                r.get("class_name", ""),
                date,
                r.get("time_in") or "",
                r.get("time_out") or "",
                present,
            ]
        )

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    config.EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def update_row_in_excel(
    school_id: int,
    school_name: str,
    record: dict[str, Any],
    date: str | None = None,
) -> Path:
    """Full rebuild is safest and fine at school scale."""
    return write_attendance_excel(school_id, school_name, date)
