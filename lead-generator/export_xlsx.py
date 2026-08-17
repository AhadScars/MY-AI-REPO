"""Excel export for leads."""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
EXPORTS = ROOT / "exports"

NAVY = "1B365D"
RED = "9B2335"
WHITE = "FFFFFF"
ALT = {
    "Manhattan": "E8EEF7",
    "Brooklyn": "EAF3EA",
    "Queens": "F7F0E4",
    "Bronx": "F3EAF6",
    "Staten Island": "E6F3F3",
}

WEBSITE_LABEL = {0: "No", 1: "Yes", 2: "Unknown"}
STATUS_LABEL = {
    "new": "New",
    "contacted": "Contacted",
    "interested": "Interested",
    "not_interested": "Not interested",
    "closed": "Closed / won",
}


def export_leads(leads: list[dict[str, Any]], filename: str | None = None) -> Path:
    EXPORTS.mkdir(parents=True, exist_ok=True)
    path = EXPORTS / (filename or f"leads_{date.today().isoformat()}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    title = Font(name="Calibri", bold=True, color=NAVY, size=18)
    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    cell_font = Font(name="Calibri", size=11, color="1F2933")
    thin = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )
    wrap = Alignment(wrap_text=True, vertical="center")
    center = Alignment(wrap_text=True, vertical="center", horizontal="center")

    ws.merge_cells("A1:N1")
    ws["A1"] = f"Lead list — {date.today().isoformat()} — {len(leads)} businesses"
    ws["A1"].font = title
    ws.row_dimensions[1].height = 26

    headers = [
        "#",
        "Business",
        "Contact",
        "Phone",
        "Email",
        "Address",
        "City / neighborhood",
        "Borough",
        "State",
        "ZIP",
        "Industry",
        "Own website?",
        "Website",
        "What appears online",
        "Status",
        "Notes",
        "Source",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(3, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[3].height = 28

    for i, lead in enumerate(leads, 1):
        row = 3 + i
        values = [
            i,
            lead.get("name"),
            lead.get("contact_name") or "",
            lead.get("phone") or "",
            lead.get("email") or "",
            lead.get("address") or "",
            lead.get("city") or lead.get("neighborhood") or "",
            lead.get("borough") or "",
            lead.get("state") or "",
            lead.get("zip") or "",
            lead.get("niche") or "",
            WEBSITE_LABEL.get(lead.get("has_website"), "Unknown"),
            lead.get("website") or "",
            lead.get("online_presence") or "",
            STATUS_LABEL.get(lead.get("status") or "new", lead.get("status") or "New"),
            lead.get("notes") or "",
            lead.get("source") or "",
        ]
        fill = PatternFill("solid", fgColor=ALT.get(lead.get("borough") or "", "FFFFFF"))
        for col, val in enumerate(values, 1):
            cell = ws.cell(row, col, val)
            cell.font = cell_font
            cell.alignment = center if col in (1, 8, 9, 10, 12, 15) else wrap
            cell.border = thin
            cell.fill = fill
            if col == 12 and lead.get("has_website") == 0:
                cell.font = Font(name="Calibri", bold=True, size=11, color=RED)
        ws.row_dimensions[row].height = 36

    widths = [5, 38, 32, 16, 24, 36, 24, 14, 8, 10, 16, 14, 32, 42, 14, 36, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    if leads:
        ws.auto_filter.ref = f"A3:Q{3 + len(leads)}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.showGridLines = False

    call = wb.create_sheet("Call sheet")
    call["A1"] = "Call sheet"
    call["A1"].font = title
    call.merge_cells("A1:E1")
    ch = ["#", "Business", "Phone", "Borough", "Talking point"]
    for col, h in enumerate(ch, 1):
        cell = call.cell(3, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
    for i, lead in enumerate(leads, 1):
        point = "No practice website found" if lead.get("has_website") == 0 else (lead.get("online_presence") or "")
        vals = [i, lead.get("name"), lead.get("phone") or "", lead.get("borough") or "", point]
        fill = PatternFill("solid", fgColor=ALT.get(lead.get("borough") or "", "FFFFFF"))
        for col, val in enumerate(vals, 1):
            cell = call.cell(3 + i, col, val)
            cell.font = cell_font
            cell.alignment = wrap
            cell.border = thin
            cell.fill = fill
        call.row_dimensions[3 + i].height = 24
    for col, w in zip("ABCDE", [5, 42, 18, 16, 42]):
        call.column_dimensions[col].width = w
    call.freeze_panes = "A4"
    call.sheet_view.showGridLines = False

    wb.save(path)
    return path
